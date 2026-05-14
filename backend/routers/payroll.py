from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional, List
from pydantic import BaseModel
import calendar, io

from database import get_db
from models import Payroll, Employee, Attendance, User, UserRole
from auth_utils import get_current_user, require_admin_hr
from payroll_engine import calc_monthly_payroll

router = APIRouter()


class PayrollIn(BaseModel):
    employee_id: int
    month:       int
    year:        int
    advance:     float = 0.0
    deposit:     float = 0.0
    incentives:  float = 0.0

class PayrollOut(BaseModel):
    id:               int
    employee_id:      int
    month:            int
    year:             int
    present_days:     float
    absent_days:      float
    total_ot_hours:   float
    total_less_hours: float
    total_out_hours:  float
    gross_salary:     float
    extra_pay:        float
    less_deduction:   float
    advance:          float
    deposit:          float
    pf_deduction:     float
    incentives:       float
    net_pay:          float
    status:           str
    class Config:
        from_attributes = True


def _get_att(db, emp_id, month, year):
    from sqlalchemy import extract
    return db.query(Attendance).filter(
        Attendance.employee_id == emp_id,
        extract("month", Attendance.date) == month,
        extract("year",  Attendance.date) == year,
    ).all()


@router.post("/calculate", response_model=PayrollOut)
def calculate(
    payload: PayrollIn,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin_hr),
):
    emp = db.query(Employee).filter(Employee.id == payload.employee_id).first()
    if not emp:
        raise HTTPException(404, "Employee not found")

    att_records = _get_att(db, payload.employee_id, payload.month, payload.year)
    result = calc_monthly_payroll(emp, att_records, payload.advance, payload.deposit, payload.incentives)

    existing = db.query(Payroll).filter(
        Payroll.employee_id == payload.employee_id,
        Payroll.month == payload.month,
        Payroll.year  == payload.year,
    ).first()

    if existing:
        for k, v in result.items():
            setattr(existing, k, v)
        existing.advance = payload.advance
        existing.deposit = payload.deposit
        existing.incentives = payload.incentives
        db.commit(); db.refresh(existing)
        return existing

    pr = Payroll(employee_id=payload.employee_id, month=payload.month, year=payload.year, **result)
    db.add(pr); db.commit(); db.refresh(pr)
    return pr


@router.get("/", response_model=List[PayrollOut])
def list_payrolls(
    month:       Optional[int] = None,
    year:        Optional[int] = None,
    employee_id: Optional[int] = None,
    db: Session  = Depends(get_db),
    user: User   = Depends(get_current_user),
):
    q = db.query(Payroll)
    if user.role == UserRole.employee:
        emp = db.query(Employee).filter(Employee.user_id == user.id).first()
        q = q.filter(Payroll.employee_id == (emp.id if emp else -1))
    elif employee_id:
        q = q.filter(Payroll.employee_id == employee_id)
    if month: q = q.filter(Payroll.month == month)
    if year:  q = q.filter(Payroll.year  == year)
    return q.all()


# ── Excel: full monthly payroll sheet ────────────────────────────────────────
@router.get("/export/monthly-excel")
def export_monthly_excel(
    month: int, year: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin_hr),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    employees = db.query(Employee).filter(Employee.status == "active").all()
    num_days  = calendar.monthrange(year, month)[1]

    def thin():
        s = Side(style="thin", color="D0D0D0")
        return Border(left=s, right=s, top=s, bottom=s)
    def fill(h): return PatternFill("solid", fgColor=h)
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    rgt = Alignment(horizontal="right", vertical="center")

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"

    # ── Summary sheet ────────────────────────────────────────────
    ws_sum.merge_cells("A1:T1")
    c = ws_sum["A1"]
    c.value = f"Monthly Payroll — {calendar.month_name[month]} {year}"
    c.font  = Font(bold=True, size=13, color="FFFFFF")
    c.fill  = fill("1E293B"); c.alignment = ctr
    ws_sum.row_dimensions[1].height = 28

    sum_hdrs = ["#","Name","Desig","Dept","Basic","Per Day","Per Hr",
                "Present","Absent","OT Hrs","Less Hrs","Gross","Extra Pay",
                "Less Ded","Advance","Deposit","PF","Incentives","Net Pay","Status"]
    for col, h in enumerate(sum_hdrs, 1):
        c = ws_sum.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=8)
        c.fill = fill("0F172A"); c.alignment = ctr; c.border = thin()
    ws_sum.row_dimensions[2].height = 30

    for idx, emp in enumerate(employees, 1):
        pr = db.query(Payroll).filter(Payroll.employee_id == emp.id,
             Payroll.month == month, Payroll.year == year).first()
        per_day = round(emp.monthly_salary / (emp.work_days_per_month or 30), 2)
        per_hr  = round(per_day / (emp.standard_hours or 8), 2)
        bg = "F8FAFC" if idx % 2 else "FFFFFF"
        row = idx + 2
        vals = [idx, emp.name, emp.designation or "", emp.department or "",
                emp.monthly_salary, per_day, per_hr,
                pr.present_days if pr else 0, pr.absent_days if pr else 0,
                pr.total_ot_hours if pr else 0, pr.total_less_hours if pr else 0,
                pr.gross_salary if pr else 0, pr.extra_pay if pr else 0,
                pr.less_deduction if pr else 0, pr.advance if pr else 0,
                pr.deposit if pr else 0, pr.pf_deduction if pr else 0,
                pr.incentives if pr else 0, pr.net_pay if pr else 0,
                pr.status.upper() if pr else "PENDING"]
        for col, val in enumerate(vals, 1):
            c = ws_sum.cell(row=row, column=col, value=val)
            c.fill = fill("D1FAE5") if col == 19 else fill(bg)
            c.border = thin(); c.alignment = ctr; c.font = Font(size=8)
            if col in [5,6,7,12,13,14,15,16,17,18,19]:
                c.number_format = "#,##0.00"

    col_ws = [4,22,8,12,10,9,9,7,7,7,7,11,9,9,9,9,8,9,11,9]
    for i, w in enumerate(col_ws, 1):
        ws_sum.column_dimensions[get_column_letter(i)].width = w
    ws_sum.freeze_panes = "A3"

    # ── Per-employee detail sheets ─────────────────────────────────
    from datetime import date as dt_date
    for emp in employees:
        ws = wb.create_sheet(title=emp.name[:28])
        pr = db.query(Payroll).filter(Payroll.employee_id == emp.id,
             Payroll.month == month, Payroll.year == year).first()
        per_day = round(emp.monthly_salary / (emp.work_days_per_month or 30), 6)
        per_hr  = round(per_day / (emp.standard_hours or 8), 6)

        ws.merge_cells("A1:AA1")
        c = ws["A1"]; c.value = f"SALARY SHEET   {emp.name.upper()}"
        c.font = Font(bold=True, size=11, color="FFFFFF")
        c.fill = fill("1E293B"); c.alignment = ctr
        ws.row_dimensions[1].height = 22

        hdrs = ["DATE","IN TIME","OUT TIME","TOTAL TIME","+","-","TIME DIFF","DAY",
                "1 FOR CAL","HOUR","BASIC","PER DAY","PER HR","PRESENT","ABSENT",
                "DUTIES","OT DUTY","OT HR","LES HR","OUT HR","TOTAL","ADV/LAB",
                "DEPOSITE","PF","EXTRA hr","LESS hr","NET PAY"]
        for col, h in enumerate(hdrs, 1):
            c = ws.cell(row=2, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF", size=7)
            c.fill = fill("0F172A"); c.alignment = ctr; c.border = thin()
        ws.row_dimensions[2].height = 28

        att_map = {a.date: a for a in _get_att(db, emp.id, month, year)}
        for day in range(1, num_days + 1):
            d   = dt_date(year, month, day)
            a   = att_map.get(d)
            row = day + 2
            is_sun = d.weekday() == 6
            bg = "FEFCE8" if is_sun else ("F8FAFC" if day % 2 else "FFFFFF")

            present   = 1 if (a and a.is_present) else 0
            absent    = 1 - present
            ot_hrs    = (a.ot_hours   if a else 0) or 0
            less_hrs  = (a.less_hours if a else 0) or 0
            out_hrs   = (a.out_hours  if a else 0) or 0
            tot_hrs   = (a.total_hours if a else 0) or 0
            t_diff    = (a.time_diff  if a else 0) or 0
            in_t      = (a.in_time    if a else "") or ""
            out_t     = (a.out_time   if a else "") or ""
            day_nm    = d.strftime("%A").upper()
            duties    = (a.duties  if a else 0) or 0
            ot_duty   = (a.ot_duty if a else 0) or 0
            col_u     = round(per_day * present, 2)
            extra_pay = round(ot_hrs * per_hr, 2)
            less_pay  = round((less_hrs + out_hrs) * per_hr, 2)
            net_day   = round(col_u + extra_pay - less_pay, 2)
            t_str     = f"{int(tot_hrs)}:{int((tot_hrs%1)*60):02d}" if tot_hrs else ""

            vals = [d.strftime("%d-%m-%Y"), in_t, out_t, t_str,
                    round(ot_hrs, 2) if ot_hrs else "",
                    round(less_hrs, 2) if less_hrs else "",
                    round(t_diff, 2), day_nm, 1, emp.standard_hours,
                    emp.monthly_salary, round(per_day, 2), round(per_hr, 2),
                    present, absent, duties, ot_duty, ot_hrs, less_hrs, out_hrs,
                    col_u, "", "", 0, extra_pay, less_pay, net_day]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.fill   = fill("BBF7D0") if col == 27 else fill(bg)
                c.border = thin(); c.alignment = ctr; c.font = Font(size=7)
                if col in [21,22,23,24,25,26,27]: c.number_format = "#,##0.00"
            ws.row_dimensions[row].height = 13

        # Summary row
        sr = num_days + 3
        if pr:
            svals = {14: pr.present_days, 15: pr.absent_days, 18: pr.total_ot_hours,
                     19: pr.total_less_hours, 20: pr.total_out_hours, 21: pr.gross_salary,
                     22: pr.advance, 23: pr.deposit, 24: pr.pf_deduction,
                     25: pr.extra_pay, 26: pr.less_deduction, 27: pr.net_pay}
            c = ws.cell(row=sr, column=1, value="TOTALS")
            c.font = Font(bold=True, size=8)
            for col, val in svals.items():
                c = ws.cell(row=sr, column=col, value=val)
                c.font = Font(bold=True, size=8, color="065F46" if col == 27 else "000000")
                c.fill = fill("E0F2FE"); c.border = thin(); c.alignment = ctr
                c.number_format = "#,##0.00"

        cw2 = [12,8,8,9,6,6,7,10,5,5,8,8,8,6,6,5,6,6,6,6,9,8,8,6,8,8,10]
        for i, w in enumerate(cw2, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A3"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"Payroll_{calendar.month_name[month]}_{year}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


# ── PDF payslip ────────────────────────────────────────────────────────────────
@router.get("/export/payslip-pdf/{employee_id}")
def export_payslip_pdf(
    employee_id: int, month: int, year: int,
    db: Session = Depends(get_db),
    user: User  = Depends(get_current_user),
):
    emp = db.query(Employee).filter(Employee.id == employee_id).first()
    if not emp: raise HTTPException(404, "Employee not found")
    pr = db.query(Payroll).filter(Payroll.employee_id == employee_id,
         Payroll.month == month, Payroll.year == year).first()
    if not pr: raise HTTPException(404, "Payroll not calculated yet for this month")

    buf = _generate_pdf(emp, pr, month, year)
    fname = f"Payslip_{emp.name.replace(' ','_')}_{calendar.month_name[month]}_{year}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


def _generate_pdf(emp, pr, month: int, year: int) -> io.BytesIO:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

    DARK   = colors.HexColor("#1E293B")
    MID    = colors.HexColor("#0F172A")
    GREEN  = colors.HexColor("#065F46")
    LGREEN = colors.HexColor("#D1FAE5")
    LRED   = colors.HexColor("#FEE2E2")
    LBLU   = colors.HexColor("#EFF6FF")
    LGREY  = colors.HexColor("#F8FAFC")
    WHITE  = colors.white

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
          topMargin=12*mm, bottomMargin=12*mm, leftMargin=14*mm, rightMargin=14*mm)

    sty = getSampleStyleSheet()
    def ps(name, **kw): return ParagraphStyle(name, parent=sty["Normal"], **kw)

    title_s  = ps("T",  fontSize=16, textColor=WHITE, alignment=TA_CENTER, fontName="Helvetica-Bold")
    sub_s    = ps("S",  fontSize=9,  textColor=WHITE, alignment=TA_CENTER)
    lbl_s    = ps("L",  fontSize=8,  textColor=DARK,  fontName="Helvetica-Bold")
    val_s    = ps("V",  fontSize=8,  textColor=DARK)
    sec_s    = ps("SE", fontSize=9,  textColor=WHITE, fontName="Helvetica-Bold", alignment=TA_CENTER)
    net_s    = ps("N",  fontSize=12, textColor=WHITE, fontName="Helvetica-Bold", alignment=TA_CENTER)
    foot_s   = ps("F",  fontSize=7,  textColor=colors.grey, alignment=TA_CENTER)

    def hdr_tbl(rows, bg=DARK):
        t = Table(rows, colWidths=[182*mm])
        t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),bg),
                               ("TOPPADDING",(0,0),(-1,-1),5),
                               ("BOTTOMPADDING",(0,0),(-1,-1),5)]))
        return t

    per_day = round(emp.monthly_salary / (emp.work_days_per_month or 30), 2)
    per_hr  = round(per_day / (emp.standard_hours or 8), 2)
    mn      = calendar.month_name[month]

    elems = []

    # Header
    elems.append(hdr_tbl([
        [Paragraph("DR. CLINIC — PAYROLL MANAGEMENT", title_s)],
        [Paragraph("SALARY SLIP", sub_s)],
        [Paragraph(f"{mn} {year}", sub_s)],
    ]))
    elems.append(Spacer(1, 5*mm))

    # Employee details table
    def row2(l1,v1,l2,v2):
        return [Paragraph(l1,lbl_s),Paragraph(str(v1),val_s),
                Paragraph(l2,lbl_s),Paragraph(str(v2),val_s)]
    emp_data = [
        row2("Employee Name", emp.name,             "Employee Code", emp.employee_code),
        row2("Designation",   emp.designation or "—","Department",  emp.department or "—"),
        row2("Basic Salary",  f"₹ {emp.monthly_salary:,.2f}", "Month/Year", f"{mn} {year}"),
        row2("Per Day",       f"₹ {per_day:,.2f}",  "Per Hour",     f"₹ {per_hr:,.2f}"),
        row2("Standard Hrs",  f"{emp.standard_hours}h/day", "Work Days", str(emp.work_days_per_month or 30)),
    ]
    emp_tbl = Table(emp_data, colWidths=[38*mm,52*mm,38*mm,52*mm])
    emp_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(0,-1),LBLU),("BACKGROUND",(2,0),(2,-1),LBLU),
        ("GRID",(0,0),(-1,-1),0.5,colors.lightgrey),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("LEFTPADDING",(0,0),(-1,-1),5),
    ]))
    elems.append(emp_tbl); elems.append(Spacer(1, 4*mm))

    # Attendance
    elems.append(hdr_tbl([[Paragraph("ATTENDANCE SUMMARY", sec_s)]], bg=MID))
    att_data = [
        row2("Present Days", pr.present_days, "Absent Days", pr.absent_days),
        row2("OT Hours",     f"{pr.total_ot_hours:.2f}","Less Hours", f"{pr.total_less_hours:.2f}"),
        row2("Out Hours",    f"{pr.total_out_hours:.2f}","Net Duty",
             str(int(pr.present_days))),
    ]
    att_tbl = Table(att_data, colWidths=[38*mm,52*mm,38*mm,52*mm])
    att_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(0,-1),LGREY),("BACKGROUND",(2,0),(2,-1),LGREY),
        ("GRID",(0,0),(-1,-1),0.5,colors.lightgrey),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("LEFTPADDING",(0,0),(-1,-1),5),
    ]))
    elems.append(att_tbl); elems.append(Spacer(1, 4*mm))

    # Earnings / Deductions
    earn_hdr = Table([[Paragraph("EARNINGS", sec_s), Spacer(3,1), Paragraph("DEDUCTIONS", sec_s)]],
                     colWidths=[89*mm, 3*mm, 89*mm])
    earn_hdr.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(0,0),GREEN),
        ("BACKGROUND",(2,0),(2,0),colors.HexColor("#991B1B")),
    ]))
    elems.append(earn_hdr)

    def money(v): return Paragraph(f"₹ {v:,.2f}", ps("MR", fontSize=8, alignment=TA_RIGHT))
    earns = [("Basic / Gross Pay", pr.gross_salary),
             ("Overtime Pay",      pr.extra_pay),
             ("Incentives",        pr.incentives)]
    deds  = [("Less Hours Deduction", pr.less_deduction),
             ("Advance / Labour",     pr.advance),
             ("PF Deduction",         pr.pf_deduction),
             ("Deposit",              pr.deposit)]
    max_r = max(len(earns), len(deds))
    ed_rows = []
    for i in range(max_r):
        el = Paragraph(earns[i][0], lbl_s) if i < len(earns) else Paragraph("", val_s)
        ev = money(earns[i][1])            if i < len(earns) else Paragraph("", val_s)
        dl = Paragraph(deds[i][0],  lbl_s) if i < len(deds)  else Paragraph("", val_s)
        dv = money(deds[i][1])             if i < len(deds)  else Paragraph("", val_s)
        ed_rows.append([el, ev, dl, dv])

    ed_tbl = Table(ed_rows, colWidths=[50*mm,39*mm,50*mm,39*mm])
    ed_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(1,-1),LGREEN),("BACKGROUND",(2,0),(3,-1),LRED),
        ("GRID",(0,0),(-1,-1),0.5,colors.lightgrey),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("LEFTPADDING",(0,0),(-1,-1),5),("RIGHTPADDING",(1,0),(1,-1),5),("RIGHTPADDING",(3,0),(3,-1),5),
    ]))
    elems.append(ed_tbl); elems.append(Spacer(1, 4*mm))

    # Net Pay
    net_tbl = Table([[Paragraph("NET PAY", net_s), Paragraph(f"₹ {pr.net_pay:,.2f}", net_s)]],
                    colWidths=[91*mm, 91*mm])
    net_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1),GREEN),
        ("TOPPADDING",(0,0),(-1,-1),9),("BOTTOMPADDING",(0,0),(-1,-1),9),
    ]))
    elems.append(net_tbl); elems.append(Spacer(1, 6*mm))
    elems.append(HRFlowable(width="100%", color=colors.lightgrey))
    elems.append(Spacer(1,2*mm))
    elems.append(Paragraph("This is a computer-generated salary slip and does not require a signature.", foot_s))

    doc.build(elems)
    buf.seek(0)
    return buf
