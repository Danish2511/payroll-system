from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import extract
import calendar, io

from database import get_db
from models import Employee, Attendance, Payroll, User
from auth_utils import require_admin_hr

router = APIRouter()


@router.get("/attendance-excel")
def attendance_excel(
    month: int, year: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin_hr),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import date as dt

    employees = db.query(Employee).filter(Employee.status == "active").order_by(Employee.name).all()
    num_days  = calendar.monthrange(year, month)[1]

    def thin():
        s = Side(style="thin", color="D0D0D0")
        return Border(left=s, right=s, top=s, bottom=s)
    def fill(h): return PatternFill("solid", fgColor=h)
    ctr = Alignment(horizontal="center", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    day_hdrs = [str(d) for d in range(1, num_days+1)]
    all_hdrs = ["#","Name","Desig","Dept"] + day_hdrs + ["Present","Absent","OT Hrs","Less Hrs"]
    ncols = len(all_hdrs)

    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]; c.value = f"Attendance Report — {calendar.month_name[month]} {year}"
    c.font = Font(bold=True, size=12, color="FFFFFF"); c.fill = fill("1E293B"); c.alignment = ctr
    ws.row_dimensions[1].height = 26

    for col, h in enumerate(all_hdrs, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=8)
        c.fill = fill("0F172A"); c.alignment = ctr; c.border = thin()
    ws.row_dimensions[2].height = 28

    for idx, emp in enumerate(employees, 1):
        row = idx + 2
        att_records = {
            a.date: a for a in
            db.query(Attendance).filter(
                Attendance.employee_id == emp.id,
            ).all()
            if a.date.month == month and a.date.year == year
        }
        ws.cell(row=row, column=1, value=idx).border = thin()
        ws.cell(row=row, column=2, value=emp.name).border = thin()
        ws.cell(row=row, column=3, value=emp.designation or "").border = thin()
        ws.cell(row=row, column=4, value=emp.department or "").border = thin()

        present = 0
        for day in range(1, num_days+1):
            d = dt(year, month, day)
            a = att_records.get(d)
            is_sun = d.weekday() == 6
            if a and a.is_present:
                val = "P"; bg = "D1FAE5"; present += 1
            elif is_sun:
                val = "H"; bg = "FEF3C7"
            else:
                val = "A"; bg = "FEE2E2"
            c = ws.cell(row=row, column=4+day, value=val)
            c.fill = fill(bg[1:]); c.border = thin()
            c.alignment = ctr; c.font = Font(size=7)

        total_ot   = sum((a.ot_hours   or 0) for a in att_records.values() if a.is_present)
        total_less = sum((a.less_hours or 0) for a in att_records.values() if a.is_present)
        off = 4 + num_days
        for col, val in [(off+1, present),(off+2, num_days-present),
                         (off+3, round(total_ot,2)),(off+4, round(total_less,2))]:
            c = ws.cell(row=row, column=col, value=val)
            c.border = thin(); c.alignment = ctr; c.font = Font(size=8)

    # Column widths
    for c in range(1, 5):
        ws.column_dimensions[get_column_letter(c)].width = [4,22,8,12][c-1]
    for c in range(5, 5+num_days):
        ws.column_dimensions[get_column_letter(c)].width = 3.2
    for c in range(5+num_days, 5+num_days+4):
        ws.column_dimensions[get_column_letter(c)].width = 8

    ws.freeze_panes = "E3"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"Attendance_{calendar.month_name[month]}_{year}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


@router.get("/summary")
def report_summary(
    month: int, year: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_admin_hr),
):
    payrolls = db.query(Payroll).filter(Payroll.month == month, Payroll.year == year).all()
    result = []
    for p in payrolls:
        emp = db.query(Employee).filter(Employee.id == p.employee_id).first()
        if emp:
            result.append({
                "name":        emp.name,
                "designation": emp.designation,
                "department":  emp.department,
                "present":     p.present_days,
                "absent":      p.absent_days,
                "ot_hours":    p.total_ot_hours,
                "gross":       p.gross_salary,
                "advance":     p.advance,
                "deposit":     p.deposit,
                "pf":          p.pf_deduction,
                "extra_pay":   p.extra_pay,
                "less_ded":    p.less_deduction,
                "net_pay":     p.net_pay,
            })
    return result
